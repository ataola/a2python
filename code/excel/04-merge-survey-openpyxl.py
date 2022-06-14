#!/usr/bin/env python
# -*- coding: UTF-8 -*-
from openpyxl import Workbook
from openpyxl import load_workbook
from pathlib import Path, PurePath

src_path = './survey'
dst_path = './result/result.xlsx'

p = Path(src_path)
files = [x for x in p.iterdir() if PurePath(x).match('*.xlsx')]

content = []

for file in files:
    username = file.stem
    wb = load_workbook(file)
    ws = wb.active
    answer1 = ws['E5'].value
    answer2 = ws.cell(row=11, column=5).value
    temp = f'{username},{answer1},{answer2}'
    content.append(temp.split(','))

table_header = ['员工姓名', '第一题', '第二题']
wb = Workbook()
ws = wb.active
ws = wb.create_sheet("统计结果")
rw = 1
col = 1
for header_field in table_header:
    ws.cell(row=rw, column=col, value=header_field)
    col += 1

rw += 1
for line in content:
    col = 1
    for cell in line:
        ws.cell(row=rw, column=col, value=cell)
        col += 1
    rw += 1

wb.save(dst_path)
